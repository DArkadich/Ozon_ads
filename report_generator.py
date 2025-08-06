"""Report generation module for Ozon advertising campaigns."""
import os
from datetime import datetime
from typing import Dict, List, Optional
import pandas as pd
from fpdf import FPDF
from jinja2 import Template
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from loguru import logger
from config import settings


class ReportGenerator:
    """Generator for campaign performance reports."""
    
    def __init__(self):
        """Initialize report generator."""
        self.output_dir = settings.report_output_dir
        os.makedirs(self.output_dir, exist_ok=True)
    
    def generate_excel_report(self, campaign_summary: Dict, keyword_analysis: List[Dict], 
                            filename: Optional[str] = None) -> str:
        """Generate Excel report with campaign analysis."""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ozon_campaign_report_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        logger.info(f"Generating Excel report: {filepath}")
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        self._create_summary_sheet(wb, campaign_summary)
        
        # Create keywords analysis sheet
        self._create_keywords_sheet(wb, keyword_analysis)
        
        # Create recommendations sheet
        self._create_recommendations_sheet(wb, campaign_summary, keyword_analysis)
        
        # Save workbook
        wb.save(filepath)
        logger.info(f"Excel report saved: {filepath}")
        
        return filepath
    
    def _create_summary_sheet(self, workbook, campaign_summary: Dict):
        """Create campaign summary sheet."""
        ws = workbook.create_sheet("Сводка кампании")
        
        # Headers styling
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        ws['A1'] = f"Отчёт по кампании {campaign_summary.get('campaign_id', 'N/A')}"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Generation date
        ws['A2'] = f"Дата создания: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws.merge_cells('A2:D2')
        
        # Performance metrics
        row = 4
        ws[f'A{row}'] = "Основные показатели"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        metrics = campaign_summary.get('performance_metrics', {})
        
        row += 1
        performance_data = [
            ("Общие расходы", f"{metrics.get('total_spend', 0):.2f} ₽"),
            ("Общая выручка", f"{metrics.get('total_revenue', 0):.2f} ₽"),
            ("Общий CTR", f"{metrics.get('overall_ctr', 0):.2f}%"),
            ("Общий CR", f"{metrics.get('overall_cr', 0):.2f}%"),
            ("Общий ДРР", f"{metrics.get('overall_drr', 0):.2f}%"),
            ("ROI", f"{metrics.get('overall_roi', 0):.2f}"),
            ("Всего кликов", f"{metrics.get('total_clicks', 0):,}"),
            ("Всего показов", f"{metrics.get('total_impressions', 0):,}"),
            ("Всего заказов", f"{metrics.get('total_orders', 0):,}")
        ]
        
        for metric_name, metric_value in performance_data:
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = metric_value
            row += 1
        
        # Actions summary
        row += 2
        ws[f'A{row}'] = "Необходимые действия"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        actions = campaign_summary.get('actions_needed', {})
        row += 1
        
        action_labels = {
            'pause': 'Отключить ключей',
            'increase_bid': 'Повысить ставки',
            'decrease_bid': 'Понизить ставки',
            'monitor': 'Мониторить'
        }
        
        for action, count in actions.items():
            if count > 0:
                ws[f'A{row}'] = action_labels.get(action, action)
                ws[f'B{row}'] = count
                row += 1
        
        # Recommendations
        recommendations = campaign_summary.get('recommendations', [])
        if recommendations:
            row += 2
            ws[f'A{row}'] = "Рекомендации"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:D{row}')
            
            for rec in recommendations:
                row += 1
                ws[f'A{row}'] = rec
                ws.merge_cells(f'A{row}:D{row}')
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column_cells:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            if column_letter:
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    def _create_keywords_sheet(self, workbook, keyword_analysis: List[Dict]):
        """Create keywords analysis sheet."""
        ws = workbook.create_sheet("Анализ ключевых слов")
        
        # Convert to DataFrame
        df = pd.DataFrame(keyword_analysis)
        
        if df.empty:
            ws['A1'] = "Нет данных для анализа"
            return
        
        # Select and order columns
        columns = [
            'keyword', 'impressions', 'clicks', 'ctr', 'orders', 'cr', 
            'spend', 'revenue', 'drr', 'action', 'recommendation', 'priority'
        ]
        
        # Filter existing columns
        available_columns = [col for col in columns if col in df.columns]
        df_filtered = df[available_columns]
        
        # Rename columns to Russian
        column_names = {
            'keyword': 'Ключевое слово',
            'impressions': 'Показы',
            'clicks': 'Клики',
            'ctr': 'CTR (%)',
            'orders': 'Заказы',
            'cr': 'CR (%)',
            'spend': 'Расходы (₽)',
            'revenue': 'Выручка (₽)',
            'drr': 'ДРР (%)',
            'action': 'Действие',
            'recommendation': 'Рекомендация',
            'priority': 'Приоритет'
        }
        
        df_filtered = df_filtered.rename(columns=column_names)
        
        # Add DataFrame to worksheet
        for r in dataframe_to_rows(df_filtered, index=False, header=True):
            ws.append(r)
        
        # Style header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Color-code action column
        action_colors = {
            'pause': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            'increase_bid': PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
            'decrease_bid': PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
            'monitor': PatternFill(start_color="A8E6CF", end_color="A8E6CF", fill_type="solid")
        }
        
        action_col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Действие':
                action_col_idx = idx
                break
        
        if action_col_idx:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, 
                                  min_col=action_col_idx, max_col=action_col_idx):
                cell = row[0]
                if cell.value in action_colors:
                    cell.fill = action_colors[cell.value]
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column_cells:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            if column_letter:
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    def _create_recommendations_sheet(self, workbook, campaign_summary: Dict, keyword_analysis: List[Dict]):
        """Create detailed recommendations sheet."""
        ws = workbook.create_sheet("Детальные рекомендации")
        
        # Title
        ws['A1'] = "Детальные рекомендации по оптимизации"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')
        
        row = 3
        
        # Critical issues
        critical_keywords = [k for k in keyword_analysis if k.get('priority', 0) >= 90]
        if critical_keywords:
            ws[f'A{row}'] = "🔴 КРИТИЧЕСКИЕ ПРОБЛЕМЫ (требуют немедленного внимания)"
            ws[f'A{row}'].font = Font(bold=True, color="FF0000")
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            for keyword in critical_keywords[:10]:  # Top 10
                ws[f'A{row}'] = keyword.get('keyword', '')
                ws[f'B{row}'] = keyword.get('recommendation', '')
                ws[f'C{row}'] = f"Приоритет: {keyword.get('priority', 0)}"
                row += 1
            row += 1
        
        # High-performance keywords
        high_performers = [k for k in keyword_analysis if 'high_performance' in k.get('issues', [])]
        if high_performers:
            ws[f'A{row}'] = "📈 ВЫСОКОЭФФЕКТИВНЫЕ КЛЮЧИ (можно масштабировать)"
            ws[f'A{row}'].font = Font(bold=True, color="008000")
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            for keyword in high_performers[:10]:
                ws[f'A{row}'] = keyword.get('keyword', '')
                ws[f'B{row}'] = keyword.get('recommendation', '')
                ws[f'C{row}'] = f"CTR: {keyword.get('ctr', 0):.2f}% | CR: {keyword.get('cr', 0):.2f}%"
                row += 1
            row += 1
        
        # Bid adjustments
        bid_adjustments = [k for k in keyword_analysis if k.get('bid_adjustment', 0) != 0]
        if bid_adjustments:
            ws[f'A{row}'] = "💰 КОРРЕКТИРОВКА СТАВОК"
            ws[f'A{row}'].font = Font(bold=True, color="0066CC")
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            for keyword in bid_adjustments[:15]:
                adjustment = keyword.get('bid_adjustment', 0)
                direction = "↗️" if adjustment > 0 else "↘️"
                ws[f'A{row}'] = keyword.get('keyword', '')
                ws[f'B{row}'] = f"{direction} {abs(adjustment):.1f}%"
                ws[f'C{row}'] = keyword.get('recommendation', '')
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)
    
    def generate_pdf_report(self, campaign_summary: Dict, filename: Optional[str] = None) -> str:
        """Generate PDF report with campaign summary."""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ozon_campaign_summary_{timestamp}.pdf"
        
        filepath = os.path.join(self.output_dir, filename)
        
        logger.info(f"Generating PDF report: {filepath}")
        
        # Create PDF
        pdf = FPDF()
        pdf.add_page()
        
        # Add UTF-8 font for Russian text
        pdf.add_font('DejaVu', '', '/System/Library/Fonts/DejaVuSans.ttf', uni=True)
        pdf.set_font('DejaVu', '', 16)
        
        # Title
        campaign_id = campaign_summary.get('campaign_id', 'N/A')
        pdf.cell(0, 10, f'Отчёт по кампании {campaign_id}', 0, 1, 'C')
        pdf.ln(5)
        
        # Date
        pdf.set_font('DejaVu', '', 12)
        pdf.cell(0, 10, f'Дата: {datetime.now().strftime("%d.%m.%Y %H:%M")}', 0, 1)
        pdf.ln(5)
        
        # Performance metrics
        pdf.set_font('DejaVu', '', 14)
        pdf.cell(0, 10, 'Основные показатели:', 0, 1)
        pdf.set_font('DejaVu', '', 12)
        
        metrics = campaign_summary.get('performance_metrics', {})
        
        metrics_text = [
            f"• Общие расходы: {metrics.get('total_spend', 0):.2f} ₽",
            f"• Общая выручка: {metrics.get('total_revenue', 0):.2f} ₽",
            f"• CTR: {metrics.get('overall_ctr', 0):.2f}%",
            f"• CR: {metrics.get('overall_cr', 0):.2f}%",
            f"• ДРР: {metrics.get('overall_drr', 0):.2f}%",
            f"• ROI: {metrics.get('overall_roi', 0):.2f}",
        ]
        
        for text in metrics_text:
            pdf.cell(0, 8, text, 0, 1)
        
        pdf.ln(5)
        
        # Recommendations
        recommendations = campaign_summary.get('recommendations', [])
        if recommendations:
            pdf.set_font('DejaVu', '', 14)
            pdf.cell(0, 10, 'Рекомендации:', 0, 1)
            pdf.set_font('DejaVu', '', 12)
            
            for rec in recommendations[:5]:  # Top 5 recommendations
                # Word wrap for long recommendations
                if len(rec) > 80:
                    words = rec.split(' ')
                    lines = []
                    current_line = ""
                    
                    for word in words:
                        if len(current_line + word) < 80:
                            current_line += word + " "
                        else:
                            lines.append(current_line.strip())
                            current_line = word + " "
                    
                    if current_line:
                        lines.append(current_line.strip())
                    
                    for line in lines:
                        pdf.cell(0, 8, f"• {line}", 0, 1)
                else:
                    pdf.cell(0, 8, f"• {rec}", 0, 1)
        
        # Save PDF
        pdf.output(filepath)
        logger.info(f"PDF report saved: {filepath}")
        
        return filepath
    
    def generate_html_report(self, campaign_summary: Dict, keyword_analysis: List[Dict],
                           filename: Optional[str] = None) -> str:
        """Generate HTML report for web viewing."""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ozon_campaign_report_{timestamp}.html"
        
        filepath = os.path.join(self.output_dir, filename)
        
        # HTML template
        template_str = """
        <!DOCTYPE html>
        <html lang="ru">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Отчёт по кампании {{ campaign_id }}</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                .header { text-align: center; margin-bottom: 30px; }
                .metrics { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 20px 0; }
                .metric-card { background: #f5f5f5; padding: 15px; border-radius: 8px; }
                .metric-value { font-size: 24px; font-weight: bold; color: #333; }
                .metric-label { color: #666; margin-top: 5px; }
                .recommendations { background: #e8f4f8; padding: 20px; border-radius: 8px; margin: 20px 0; }
                .critical { background: #ffe6e6; border-left: 4px solid #ff4444; }
                .success { background: #e6ffe6; border-left: 4px solid #44ff44; }
                .warning { background: #fff3e0; border-left: 4px solid #ff9800; }
                table { width: 100%; border-collapse: collapse; margin: 20px 0; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                th { background-color: #f2f2f2; }
                .action-pause { background-color: #ffebee; }
                .action-increase { background-color: #e8f5e8; }
                .action-decrease { background-color: #fff3e0; }
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Отчёт по кампании {{ campaign_id }}</h1>
                <p>Дата создания: {{ date }}</p>
            </div>
            
            <div class="metrics">
                <div class="metric-card">
                    <div class="metric-value">{{ metrics.total_spend|round(2) }} ₽</div>
                    <div class="metric-label">Общие расходы</div>
                </div>
                <div class="metric-card">
                    <div class="metric-value">{{ metrics.overall_ctr|round(2) }}%</div>
                    <div class="metric-label">CTR</div>
                </div>
                <div class="metric-card">
                    <div class="metric-value">{{ metrics.overall_cr|round(2) }}%</div>
                    <div class="metric-label">CR</div>
                </div>
                <div class="metric-card">
                    <div class="metric-value">{{ metrics.overall_drr|round(2) }}%</div>
                    <div class="metric-label">ДРР</div>
                </div>
            </div>
            
            {% if recommendations %}
            <div class="recommendations">
                <h2>Рекомендации</h2>
                <ul>
                {% for rec in recommendations %}
                    <li>{{ rec }}</li>
                {% endfor %}
                </ul>
            </div>
            {% endif %}
            
            {% if critical_keywords %}
            <div class="critical">
                <h2>🔴 Критические проблемы</h2>
                <ul>
                {% for keyword in critical_keywords[:10] %}
                    <li><strong>{{ keyword.keyword }}</strong>: {{ keyword.recommendation }}</li>
                {% endfor %}
                </ul>
            </div>
            {% endif %}
        </body>
        </html>
        """
        
        template = Template(template_str)
        
        # Prepare data
        critical_keywords = [k for k in keyword_analysis if k.get('priority', 0) >= 90]
        
        html_content = template.render(
            campaign_id=campaign_summary.get('campaign_id', 'N/A'),
            date=datetime.now().strftime('%d.%m.%Y %H:%M'),
            metrics=campaign_summary.get('performance_metrics', {}),
            recommendations=campaign_summary.get('recommendations', []),
            critical_keywords=critical_keywords
        )
        
        # Save HTML
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        logger.info(f"HTML report saved: {filepath}")
        return filepath